from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EN_HEADERS = [
    "Source Period",
    "Period Start Date",
    "Period End Date",
    "Total Payable",
    "Currency",
    "Transaction Key",
    "Transaction Posted Timestamp",
    "Year",
    "Month",
    "Quarter",
    "Transaction Type",
    "Transaction Type(中文)",
    "Transaction Description",
    "Transaction Description(中文)",
    "Customer Order #",
    "Customer Order line #",
    "Purchase Order #",
    "Purchase Order line #",
    "Amount",
    "Amount in CNY",
    "Exchange Rate",
    "Amount Type",
    "Amount Type(中文)",
    "Data Type Description",
    "Ship Qty",
    "Commission Rate",
    "Transaction Reason Description",
    "Partner Item Id",
    "Partner GTIN",
    "Partner Item Name",
    "Product Tax Code",
    "Ship to State",
    "Ship to City",
    "Ship to Zipcode",
    "Contract Category",
    "Product Type",
    "Commission Rule",
    "Commission Rule(中文)",
    "Shipping Method",
    "Fulfillment Type",
    "Fulfillment Type(中文)",
    "Fulfillment Details",
    "Original Commission",
    "Commission Incentive Program",
    "Commission Saving",
    "Customer Promo Type",
    "Total Walmart Funded Savings Program",
    "Campaign Id",
    "Item Condition",
    "Item Condition(中文)",
]

CN_HEADERS = [
    "来源账期",
    "账期开始日期",
    "账期结束日期",
    "应付总金额",
    "货币",
    "交易键",
    "交易发布时间",
    "年份",
    "月份",
    "季度",
    "交易类型",
    "交易类型(中文)",
    "交易描述",
    "交易描述(中文)",
    "客户订单号",
    "客户订单行号",
    "采购订单号",
    "采购订单行号",
    "金额",
    "人民币金额",
    "汇率(USD/CNY)",
    "金额类型",
    "金额类型(中文)",
    "数据分析类型",
    "发货数量",
    "佣金费率",
    "交易原因描述",
    "合作伙伴商品ID",
    "合作伙伴GTIN",
    "商品名称",
    "商品税码",
    "收货州",
    "收货城市",
    "收货邮编",
    "合同类别",
    "商品类型",
    "佣金规则",
    "佣金规则(中文)",
    "配送方式",
    "履单类型",
    "履单类型(中文)",
    "履单详情",
    "原始佣金",
    "佣金激励计划",
    "佣金节省",
    "客户促销类型",
    "沃尔玛资助节省计划总额",
    "活动ID",
    "商品状况",
    "商品状况(中文)",
]

TARGET_SOURCE_HEADERS = [
    header
    for header in EN_HEADERS
    if header
    not in {
        "Source Period",
        "Year",
        "Month",
        "Quarter",
        "Transaction Type(中文)",
        "Transaction Description(中文)",
        "Amount in CNY",
        "Exchange Rate",
        "Amount Type(中文)",
        "Data Type Description",
        "Commission Rule(中文)",
        "Fulfillment Type(中文)",
        "Item Condition(中文)",
    }
]

TRANSACTION_TYPE_CN = {
    "Sale": "销售",
    "Adjustment": "调整",
    "Refund": "退款",
    "Service Fee": "服务费",
    "Campaigns": "营销活动",
    "PaymentSummary": "付款汇总",
    "Reserve": "预留资金",
    "Release Reserve": "释放预留资金",
}

TRANSACTION_DESCRIPTION_CN = {
    "Purchase": "购买/销售",
    "WFS Fulfillment fee": "WFS履单费",
    "WFS Charge": "WFS费用",
    "Return Refund": "退货退款",
    "SEM Marketing": "SEM营销费",
    "WFS StorageFee": "WFS仓储费",
    "Walmart Product Advertising": "沃尔玛商品广告费",
    "WFS Return Processing Fee": "WFS退货处理费",
    "Deposited in PINGPONG account": "已存入PINGPONG账户",
    "WFS RC_InventoryDisposalFee": "WFS库存销毁费",
    "Review Accelerator": "评价加速计划费",
    "Seller Triggered Refund": "卖家主动退款",
    "WMSC Logistics Services": "沃尔玛供应链物流服务",
    "WFS LostInventory": "WFS库存丢失赔偿",
    '"WFS Return Shipping fee "': '"WFS Return Shipping fee "',
    "WFS Return Shipping fee": "WFS退货运费",
    "Business initiated discrete partner credit": "平台主动合作伙伴信用补偿",
    "WFS Refund": "WFS退款",
    "WFS FoundInventory": "WFS库存找回",
    "WFS InventoryTransferFee": "WFS库存转移费",
    "Misc Invoices": "杂项发票",
    "Funds withheld to cover future refunds": "资金扣留(备付未来退款)",
    "Reserve credited back": "预留资金退回",
}

AMOUNT_TYPE_CN = {
    "Fee/Reimbursement": "费用/补偿",
    "WFS Fee/Reimbursement": "WFS费/补偿",
    "Product Price": "商品售价",
    "Commission on Product": "商品佣金",
    "Product tax": "商品税",
    "Product tax withheld": "预扣商品税",
    "SEM Marketing Fee": "SEM营销费",
    "Total Walmart Funded Savings": "沃尔玛资助节省总额",
    "WFS Inventory Fee/Reimbursement": "WFS库存费/补偿",
    "Review Accelerator Fee": "评价加速计划费",
    "Other tax (Fees)": "其他税费",
    "Ocean freight and related services": "海运及相关服务费",
    "Promo Code": "促销码折扣",
    "Extra Savings": "额外折扣",
}

DATA_TYPE_BY_AMOUNT = {
    "Product Price": "商品售价",
    "Commission on Product": "平台佣金",
    "Product tax": "商品税(代收)",
    "Product tax withheld": "商品税(预扣)",
    "SEM Marketing Fee": "SEM营销费",
    "Total Walmart Funded Savings": "沃尔玛资助折扣",
    "WFS Fee/Reimbursement": "WFS费/补偿",
    "Review Accelerator Fee": "评价加速计划费",
    "Other tax (Fees)": "其他税费",
    "Ocean freight and related services": "海运及相关服务",
    "Promo Code": "促销码折扣",
    "Extra Savings": "额外折扣",
}

DATA_TYPE_BY_DESCRIPTION = {
    "WFS Fulfillment fee": "WFS履单费",
    "WFS Charge": "WFS费用",
    "WFS StorageFee": "WFS仓储费",
    "Walmart Product Advertising": "商品广告费",
    "WFS Return Processing Fee": "WFS退货处理费",
    "Deposited in PINGPONG account": "平台付款入账",
    "WFS RC_InventoryDisposalFee": "WFS库存销毁费",
    "WFS LostInventory": "WFS库存丢失赔偿",
    "WFS Return Shipping fee": "WFS退货运费",
    '"WFS Return Shipping fee "': "WFS退货运费",
    "Business initiated discrete partner credit": "合作伙伴信用补偿",
    "WFS Refund": "WFS库存费/补偿",
    "WFS FoundInventory": "WFS库存找回调整",
    "WFS InventoryTransferFee": "WFS库存转移费",
    "Misc Invoices": "杂项发票",
    "Funds withheld to cover future refunds": "预留金扣留",
    "Reserve credited back": "预留金释放",
}

MONTHLY_USD_CNY_RATES = {
    (2025, 5): 7.1848,
    (2025, 6): 7.1586,
    (2025, 7): 7.1494,
    (2025, 8): 7.1030,
    (2025, 9): 7.1055,
    (2025, 10): 7.0880,
    (2025, 11): 7.0789,
    (2025, 12): 6.9955,
    (2026, 1): 6.9518,
    (2026, 2): 6.8582,
    (2026, 3): 6.8947,
    (2026, 4): 6.8282,
    (2026, 5): 6.7664,
}


@dataclass
class WalmartTransactionJob:
    job_id: str
    source_label: str
    output_dir: Path
    total_path: Path
    audit_path: Path
    summary: dict
    rows: list[dict]


def process_walmart_transaction_folder(folder: Path, output_dir: Path, job_id: str, label: str | None = None) -> WalmartTransactionJob:
    label = label or folder.name
    output_dir.mkdir(parents=True, exist_ok=True)
    source_files = sorted(path for path in folder.rglob("*") if path.is_file())
    workbook_files = [path for path in source_files if path.suffix.lower() in {".xlsx", ".xlsm"} and not _is_generated_file(path)]
    skipped_files = [
        str(path.relative_to(folder))
        for path in source_files
        if path.suffix.lower() not in {".xlsx", ".xlsm"} or _is_generated_file(path)
    ]

    output_rows: list[list[Any]] = []
    file_rows: list[dict] = []
    unmapped = {
        "transaction_type": Counter(),
        "transaction_description": Counter(),
        "amount_type": Counter(),
    }
    extra_columns: Counter[str] = Counter()
    warnings = 0

    for path in workbook_files:
        parsed = _parse_walmart_workbook(path, folder, unmapped, extra_columns)
        output_rows.extend(parsed["rows"])
        file_rows.append(parsed["file_row"])
        if parsed["file_row"]["status"] != "通过":
            warnings += 1

    total_path = output_dir / f"沃尔玛经营数据-{_output_period_label(output_rows, job_id)}.xlsx"
    audit_path = output_dir / f"沃尔玛清洗审计-{job_id}.xlsx"
    _write_result_workbook(total_path, output_rows)
    _write_audit_workbook(audit_path, file_rows, skipped_files, unmapped, extra_columns)

    total_amount = sum((row[EN_HEADERS.index("Amount")] or 0) for row in output_rows if isinstance(row[EN_HEADERS.index("Amount")], (int, float)))
    total_amount_cny = sum((row[EN_HEADERS.index("Amount in CNY")] or 0) for row in output_rows if isinstance(row[EN_HEADERS.index("Amount in CNY")], (int, float)))
    source_periods = sorted({row[0] for row in output_rows if row and row[0]})
    summary = {
        "source_files": len(workbook_files),
        "skipped_files": len(skipped_files),
        "source_rows": sum(item.get("source_rows", 0) for item in file_rows),
        "parsed_rows": len(output_rows),
        "output_rows": len(output_rows),
        "source_periods": len(source_periods),
        "period_min": source_periods[0] if source_periods else "",
        "period_max": source_periods[-1] if source_periods else "",
        "total_amount": round(total_amount, 2),
        "total_amount_cny": round(total_amount_cny, 2),
        "warnings": warnings + len(skipped_files),
        "unmapped_values": sum(sum(counter.values()) for counter in unmapped.values()),
        "extra_columns": len(extra_columns),
        "output_filename": total_path.name,
        "audit_filename": audit_path.name,
    }
    rows = [
        {
            "source_file": item["source_file"],
            "sheet": item["sheet"],
            "source_period": item["source_period"],
            "source_rows": item["source_rows"],
            "parsed_rows": item["parsed_rows"],
            "date_min": item["date_min"],
            "date_max": item["date_max"],
            "status": item["status"],
            "notes": item["notes"],
        }
        for item in file_rows
    ]
    return WalmartTransactionJob(
        job_id=job_id,
        source_label=str(folder),
        output_dir=output_dir,
        total_path=total_path,
        audit_path=audit_path,
        summary=summary,
        rows=rows,
    )


def _parse_walmart_workbook(path: Path, root: Path, unmapped: dict[str, Counter], extra_columns: Counter[str]) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[list[Any]] = []
    sheet_names: list[str] = []
    notes: list[str] = []
    source_rows = 0
    parsed_dates: list[datetime] = []
    source_period = _source_period_from_filename(path.name)

    try:
        for ws in wb.worksheets:
            header = [clean_cell(value) for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            if not _looks_like_walmart_header(header):
                notes.append(f"{ws.title} 未识别为沃尔玛账单明细表头")
                continue
            sheet_names.append(ws.title)
            header_index = {name: index for index, name in enumerate(header) if name}
            for name in header:
                if name and name not in TARGET_SOURCE_HEADERS:
                    extra_columns[name] += 1
            for raw in ws.iter_rows(min_row=3, values_only=True):
                if not any(value is not None and str(value).strip() for value in raw):
                    continue
                source_rows += 1
                record = {name: raw[index] if index < len(raw) else None for name, index in header_index.items()}
                row = _build_output_row(source_period, record, unmapped)
                rows.append(row)
                posted = _parse_date(record.get("Transaction Posted Timestamp"))
                if posted:
                    parsed_dates.append(posted)
    finally:
        wb.close()

    if not sheet_names:
        notes.append("未找到有效账单明细工作表")
    if source_rows != len(rows):
        notes.append(f"源行数 {source_rows} / 解析行数 {len(rows)} 不一致")
    if not source_period:
        notes.append("文件名未识别来源账期")

    relative = str(path.relative_to(root))
    return {
        "rows": rows,
        "file_row": {
            "source_file": relative,
            "sheet": "、".join(sheet_names),
            "source_period": source_period,
            "source_rows": source_rows,
            "parsed_rows": len(rows),
            "date_min": min(parsed_dates).strftime("%Y-%m-%d") if parsed_dates else "",
            "date_max": max(parsed_dates).strftime("%Y-%m-%d") if parsed_dates else "",
            "status": "需复核" if notes else "通过",
            "notes": "；".join(notes),
        },
    }


def _build_output_row(source_period: str, record: dict[str, Any], unmapped: dict[str, Counter]) -> list[Any]:
    posted = _parse_date(record.get("Transaction Posted Timestamp"))
    year = posted.year if posted else None
    month = posted.month if posted else None
    rate = MONTHLY_USD_CNY_RATES.get((year, month)) if year and month else None
    amount = _to_number(record.get("Amount"))
    amount_cny = round(amount * rate, 2) if amount is not None and rate is not None else None
    transaction_type = clean_cell(record.get("Transaction Type"))
    transaction_description = clean_cell(record.get("Transaction Description"))
    amount_type = clean_cell(record.get("Amount Type"))
    transaction_type_cn = _translate(transaction_type, TRANSACTION_TYPE_CN, unmapped["transaction_type"])
    transaction_description_cn = _translate(transaction_description, TRANSACTION_DESCRIPTION_CN, unmapped["transaction_description"])
    amount_type_cn = _translate(amount_type, AMOUNT_TYPE_CN, unmapped["amount_type"])
    output: dict[str, Any] = {
        "Source Period": source_period,
        "Year": year,
        "Month": month,
        "Quarter": _quarter_for(month),
        "Transaction Type(中文)": transaction_type_cn,
        "Transaction Description(中文)": transaction_description_cn,
        "Amount in CNY": amount_cny,
        "Exchange Rate": rate,
        "Amount Type(中文)": amount_type_cn,
        "Data Type Description": _classify_data_type(transaction_type, transaction_description, amount_type),
        "Commission Rule(中文)": clean_cell(record.get("Commission Rule")) or None,
        "Fulfillment Type(中文)": clean_cell(record.get("Fulfillment Type")) or None,
        "Item Condition(中文)": _normalize_dash(record.get("Item Condition")),
    }
    for header in TARGET_SOURCE_HEADERS:
        value = record.get(header)
        if header == "Amount":
            value = amount
        elif header in {"Total Payable", "Ship Qty", "Commission Rate", "Original Commission", "Commission Saving", "Total Walmart Funded Savings Program"}:
            value = _to_number(value)
        elif header in {"Customer Order line #", "Purchase Order line #"}:
            value = _to_int(value)
        else:
            value = clean_cell(value) or None
        output[header] = value
    return [output.get(header) for header in EN_HEADERS]


def _write_result_workbook(path: Path, rows: list[list[Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "账单明细"
    ws.append(EN_HEADERS)
    ws.append(CN_HEADERS)
    for row in rows:
        ws.append(row)
    _style_main_sheet(ws)
    wb.save(path)


def _write_audit_workbook(
    path: Path,
    file_rows: list[dict],
    skipped_files: list[str],
    unmapped: dict[str, Counter],
    extra_columns: Counter[str],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "文件审计"
    headers = ["来源文件", "工作表", "来源账期", "源行数", "解析行数", "日期最小值", "日期最大值", "状态", "备注"]
    ws.append(headers)
    for item in file_rows:
        ws.append([
            item["source_file"],
            item["sheet"],
            item["source_period"],
            item["source_rows"],
            item["parsed_rows"],
            item["date_min"],
            item["date_max"],
            item["status"],
            item["notes"],
        ])
    _style_audit_sheet(ws)

    skipped_ws = wb.create_sheet("跳过文件")
    skipped_ws.append(["文件路径", "原因"])
    for item in skipped_files:
        skipped_ws.append([item, "非 xlsx/xlsm 或生成结果文件"])
    _style_audit_sheet(skipped_ws)

    unmapped_ws = wb.create_sheet("未映射值")
    unmapped_ws.append(["字段", "原始值", "出现次数"])
    for field, counter in unmapped.items():
        for value, count in counter.most_common():
            unmapped_ws.append([field, value, count])
    _style_audit_sheet(unmapped_ws)

    extra_ws = wb.create_sheet("源文件新增列")
    extra_ws.append(["源字段", "出现文件数"])
    for value, count in extra_columns.most_common():
        extra_ws.append([value, count])
    _style_audit_sheet(extra_ws)
    wb.save(path)


def _style_main_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="111827")
    sub_fill = PatternFill("solid", fgColor="EAF7EF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[2]:
        cell.fill = sub_fill
        cell.font = Font(color="111827", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 18,
        "B": 16,
        "C": 16,
        "D": 14,
        "G": 18,
        "K": 18,
        "M": 28,
        "N": 24,
        "T": 14,
        "X": 22,
        "AD": 28,
    }
    for index in range(1, ws.max_column + 1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = widths.get(letter, 16)
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
    for col in ["D", "S", "T", "U", "Z", "AQ", "AS", "AW"]:
        for cell in ws[col][2:]:
            cell.number_format = "#,##0.00"


def _style_audit_sheet(ws) -> None:
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 22


def _translate(value: str, mapping: dict[str, str], unmapped_counter: Counter) -> str | None:
    if not value:
        return None
    normalized = _normalize_description(value)
    if normalized in mapping:
        return mapping[normalized]
    if value in mapping:
        return mapping[value]
    unmapped_counter[value] += 1
    return value


def _classify_data_type(transaction_type: str, description: str, amount_type: str) -> str | None:
    if amount_type in DATA_TYPE_BY_AMOUNT:
        return DATA_TYPE_BY_AMOUNT[amount_type]
    normalized_description = _normalize_description(description)
    if normalized_description in DATA_TYPE_BY_DESCRIPTION:
        return DATA_TYPE_BY_DESCRIPTION[normalized_description]
    if description in DATA_TYPE_BY_DESCRIPTION:
        return DATA_TYPE_BY_DESCRIPTION[description]
    if transaction_type == "PaymentSummary":
        return "平台付款入账"
    if transaction_type == "Reserve":
        return "预留金扣留"
    if transaction_type == "Release Reserve":
        return "预留金释放"
    return AMOUNT_TYPE_CN.get(amount_type) or TRANSACTION_DESCRIPTION_CN.get(description) or description or amount_type or transaction_type or None


def _looks_like_walmart_header(headers: list[str]) -> bool:
    required = {"Transaction Posted Timestamp", "Transaction Type", "Transaction Description"}
    return required.issubset(set(headers))


def _source_period_from_filename(filename: str) -> str:
    match = re.search(r"(20\d{6})[-_－—](20\d{6})", filename)
    return f"{match.group(1)}-{match.group(2)}" if match else ""


def _output_period_label(rows: list[list[Any]], fallback: str) -> str:
    periods = [row[0] for row in rows if row and row[0]]
    if periods:
        return str(max(periods)).split("-")[-1][:6]
    return fallback


def _parse_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    raw = clean_cell(value)
    if not raw:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y/%m/%d", "%m/%d/%y"):
        try:
            return datetime.strptime(raw[:10], fmt)
        except ValueError:
            pass
    return None


def _quarter_for(month: int | None) -> str | None:
    if not month:
        return None
    return f"Q{((month - 1) // 3) + 1}"


def _to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    raw = clean_cell(value).replace(",", "")
    if raw in {"", "-"}:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _to_int(value: Any) -> int | None:
    number = _to_number(value)
    return int(number) if number is not None and number == int(number) else number


def _normalize_dash(value: Any) -> str | None:
    raw = clean_cell(value)
    if not raw:
        return None
    return "-" if raw.replace(" ", "") == "-" else raw


def _normalize_description(value: str) -> str:
    return clean_cell(value).replace('"', "").strip()


def _is_generated_file(path: Path) -> bool:
    name = path.name
    return name.startswith("~$") or name.startswith("沃尔玛经营数据-") or name.startswith("沃尔玛清洗审计-")


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    return " ".join(str(value).strip().split())
