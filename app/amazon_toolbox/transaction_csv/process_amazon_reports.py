#!/usr/bin/env python3
import csv
import json
import os
import re
import shutil
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .analyze_amazon_reports import (
    MONTHS,
    ROOT,
    SKIP_DIRS,
    clean_cell,
    find_header,
    norm_header,
    parse_date,
    rows_from_csv,
)

DETAIL_DIR = ROOT / "detail_backup"
OUTPUT_DIR = ROOT / "outputs"
COUNTRY_DIR = OUTPUT_DIR / "country_merged"
MERGED_QUARTERLY_DIR = OUTPUT_DIR / "merged_quarterly_backup"
REFERENCE_RULE_PATH = Path(
    os.environ.get(
        "AMAZON_RECORD_TYPE_REFERENCE",
        "/Users/xukeqiang/Desktop/记录类型修正参考表.xlsx",
    )
)

BASE_LEFT = [
    "country",
    "source file",
    "date/time",
    "date",
    "year",
    "month",
    "record type",
    "settlement id",
    "type",
    "order id",
    "sku",
    "description",
]

TOTAL_COL = "total"

PREFERRED_ORDER = BASE_LEFT + [
    "quantity",
    "marketplace",
    "account type",
    "fulfillment",
    "order city",
    "order state",
    "order postal",
    "tax collection model",
    "product sales",
    "product sales tax",
    "shipping credits",
    "shipping credits tax",
    "gift wrap credits",
    "gift wrap credits tax",
    "regulatory fee",
    "tax on regulatory fee",
    "promotional rebates",
    "promotional rebates tax",
    "sales tax collected",
    "marketplace facilitator tax",
    "marketplace withheld tax",
    "selling fees",
    "fba fees",
    "other transaction fees",
    "other",
    "transaction status",
    "transaction release date",
    TOTAL_COL,
]

CHINESE_HEADERS = {
    "country": "国家",
    "source file": "源文件",
    "date/time": "日期时间",
    "date": "日期",
    "year": "年份",
    "month": "月份",
    "record type": "记录类型",
    "settlement id": "结算编号",
    "type": "类型",
    "order id": "订单编号",
    "sku": "SKU",
    "description": "描述",
    "quantity": "数量",
    "marketplace": "站点",
    "account type": "账户类型",
    "fulfillment": "配送方式",
    "order city": "订单城市",
    "order state": "订单州/省",
    "order postal": "订单邮编",
    "tax collection model": "税费征收模式",
    "product sales": "商品销售额",
    "product sales tax": "商品销售税",
    "shipping credits": "运费收入",
    "shipping credits tax": "运费税",
    "gift wrap credits": "礼品包装收入",
    "gift wrap credits tax": "礼品包装税",
    "regulatory fee": "监管费",
    "tax on regulatory fee": "监管费税",
    "promotional rebates": "促销折扣",
    "promotional rebates tax": "促销折扣税",
    "sales tax collected": "已收销售税",
    "marketplace facilitator tax": "平台代扣税",
    "marketplace withheld tax": "商城预扣税",
    "selling fees": "销售费用",
    "fba fees": "FBA费用",
    "other transaction fees": "其他交易费用",
    "other": "其他",
    "transaction status": "交易状态",
    "transaction release date": "交易释放日期",
    "total": "总计",
}

CANONICAL_ALIASES = {
    "date/time": "date/time",
    "date/heure": "date/time",
    "fecha/hora": "date/time",
    "fecha y hora": "date/time",
    "datum/uhrzeit": "date/time",
    "data/ora": "date/time",
    "datum/tijd": "date/time",
    "datum/tid": "date/time",
    "data/godzina": "date/time",
    "settlement id": "settlement id",
    "numéro de versement": "settlement id",
    "identifiant du paiement": "settlement id",
    "id. de liquidación": "settlement id",
    "identificador de pago": "settlement id",
    "abrechnungsnummer": "settlement id",
    "numero pagamento": "settlement id",
    "schikkings-id": "settlement id",
    "reglerings-id": "settlement id",
    "identyfikator rozliczenia": "settlement id",
    "type": "type",
    "tipo": "type",
    "typ": "type",
    "order id": "order id",
    "numéro de la commande": "order id",
    "numéro commande": "order id",
    "id. del pedido": "order id",
    "número de pedido": "order id",
    "bestellnummer": "order id",
    "numero ordine": "order id",
    "bestelnummer": "order id",
    "beställnings-id": "order id",
    "identyfikator zamówienia": "order id",
    "sku": "sku",
    "description": "description",
    "descripción": "description",
    "beschreibung": "description",
    "descrizione": "description",
    "beschrijving": "description",
    "beskrivning": "description",
    "opis": "description",
    "quantity": "quantity",
    "quantité": "quantity",
    "cantidad": "quantity",
    "menge": "quantity",
    "quantità": "quantity",
    "aantal": "quantity",
    "antal": "quantity",
    "ilość": "quantity",
    "marketplace": "marketplace",
    "site de vente": "marketplace",
    "web de amazon": "marketplace",
    "marketplace": "marketplace",
    "rynek": "marketplace",
    "marknadsplats": "marketplace",
    "account type": "account type",
    "fulfillment": "fulfillment",
    "fulfilment": "fulfillment",
    "traitement": "fulfillment",
    "expédition": "fulfillment",
    "cumplimiento": "fulfillment",
    "versand": "fulfillment",
    "gestione": "fulfillment",
    "realizzazione": "fulfillment",
    "levering": "fulfillment",
    "leverans": "fulfillment",
    "realizacja": "fulfillment",
    "order city": "order city",
    "ort der bestellung": "order city",
    "ville d'où provient la commande": "order city",
    "ville de la commande": "order city",
    "ciudad del pedido": "order city",
    "stadt der bestellung": "order city",
    "città ordine": "order city",
    "città di provenienza dell'ordine": "order city",
    "plaats bestelling": "order city",
    "stad för beställning": "order city",
    "miejscowość zamówienia": "order city",
    "order state": "order state",
    "bundesland": "order state",
    "région d'où provient la commande": "order state",
    "région de la commande": "order state",
    "état de la commande": "order state",
    "estado del pedido": "order state",
    "bundesland der bestellung": "order state",
    "provincia ordine": "order state",
    "provincia di provenienza dell'ordine": "order state",
    "staat bestelling": "order state",
    "delstat för beställning": "order state",
    "stan zamówienia": "order state",
    "order postal": "order postal",
    "postleitzahl": "order postal",
    "code postal de la commande": "order postal",
    "commande postale": "order postal",
    "código postal del pedido": "order postal",
    "postleitzahl der bestellung": "order postal",
    "codice postale ordine": "order postal",
    "cap dell'ordine": "order postal",
    "postcode bestelling": "order postal",
    "postadress för beställning": "order postal",
    "przekaz pocztowy": "order postal",
    "tax collection model": "tax collection model",
    "modèle de perception des taxes": "tax collection model",
    "modelo de recaudación de impuestos": "tax collection model",
    "steuererhebungsmodell": "tax collection model",
    "modello di riscossione delle imposte": "tax collection model",
    "product sales": "product sales",
    "umsätze": "product sales",
    "vendite": "product sales",
    "ventes de produits": "product sales",
    "ventas de productos": "product sales",
    "produktverkäufe": "product sales",
    "vendite prodotti": "product sales",
    "productverkoop": "product sales",
    "försäljning av produkter": "product sales",
    "sprzedaż produktów": "product sales",
    "product sales tax": "product sales tax",
    "produktumsatzsteuer": "product sales tax",
    "imposta sulle vendite dei prodotti": "product sales tax",
    "taxes sur la vente des produits": "product sales tax",
    "impuesto de ventas de productos": "product sales tax",
    "umsatzsteuer auf produktverkäufe": "product sales tax",
    "imposte vendita prodotti": "product sales tax",
    "shipping credits": "shipping credits",
    "postage credits": "shipping credits",
    "crédits d'expédition": "shipping credits",
    "crédits d’expédition": "shipping credits",
    "créditos de envío": "shipping credits",
    "gutschriften für versandkosten": "shipping credits",
    "gutschrift für versandkosten": "shipping credits",
    "accrediti spedizione": "shipping credits",
    "accrediti per le spedizioni": "shipping credits",
    "verzendtegoed": "shipping credits",
    "fraktkrediter": "shipping credits",
    "noty kredytowe za wysyłkę": "shipping credits",
    "shipping credits tax": "shipping credits tax",
    "taxe sur les crédits d’expédition": "shipping credits tax",
    "impuesto de abono de envío": "shipping credits tax",
    "steuer auf versandgutschrift": "shipping credits tax",
    "imposta accrediti per le spedizioni": "shipping credits tax",
    "gift wrap credits": "gift wrap credits",
    "giftwrap credits": "gift wrap credits",
    "crédits sur l'emballage cadeau": "gift wrap credits",
    "crédits d’emballage-cadeau": "gift wrap credits",
    "créditos por envoltorio de regalo": "gift wrap credits",
    "gutschriften für geschenkverpackung": "gift wrap credits",
    "gutschrift für geschenkverpackung": "gift wrap credits",
    "accrediti confezione regalo": "gift wrap credits",
    "accrediti per confezioni regalo": "gift wrap credits",
    "cadeauverpakking tegoed": "gift wrap credits",
    "krediter för presentinslagning": "gift wrap credits",
    "środki na pokrycie pakowania na prezent": "gift wrap credits",
    "gift wrap credits tax": "gift wrap credits tax",
    "giftwrap credits tax": "gift wrap credits tax",
    "taxes sur les crédits cadeaux": "gift wrap credits tax",
    "impuesto de créditos de envoltura": "gift wrap credits tax",
    "steuer auf geschenkverpackungsgutschriften": "gift wrap credits tax",
    "imposta sui crediti confezione regalo": "gift wrap credits tax",
    "regulatory fee": "regulatory fee",
    "tarifa reglamentaria": "regulatory fee",
    "tax on regulatory fee": "tax on regulatory fee",
    "impuesto sobre tarifa reglamentaria": "tax on regulatory fee",
    "promotional rebates": "promotional rebates",
    "rabais promotionnels": "promotional rebates",
    "total des réductions": "promotional rebates",
    "descuentos promocionales": "promotional rebates",
    "rebajas promocionales": "promotional rebates",
    "aktionsrabatte": "promotional rebates",
    "rabatte aus werbeaktionen": "promotional rebates",
    "sconti promozionali": "promotional rebates",
    "promotiekortingen": "promotional rebates",
    "kampanjrabatter": "promotional rebates",
    "rabaty promocyjne": "promotional rebates",
    "promotional rebates tax": "promotional rebates tax",
    "taxes sur les remises promotionnelles": "promotional rebates tax",
    "impuesto de reembolsos promocionales": "promotional rebates tax",
    "steuer auf aktionsrabatte": "promotional rebates tax",
    "imposta sugli sconti promozionali": "promotional rebates tax",
    "sales tax collected": "sales tax collected",
    "taxe de ventes prélevée": "sales tax collected",
    "inkasserad moms": "sales tax collected",
    "pobrany podatek od sprzedaży": "sales tax collected",
    "marketplace facilitator tax": "marketplace facilitator tax",
    "taxe marketplace facilitator": "marketplace facilitator tax",
    "skatt för marknadsplatsförmedlare": "marketplace facilitator tax",
    "podatek od transakcji marketplace facilitator": "marketplace facilitator tax",
    "marketplace withheld tax": "marketplace withheld tax",
    "taxes retenues sur le site de vente": "marketplace withheld tax",
    "impuesto de retenciones en la plataforma": "marketplace withheld tax",
    "einbehaltene steuer auf marketplace": "marketplace withheld tax",
    "trattenuta iva del marketplace": "marketplace withheld tax",
    "selling fees": "selling fees",
    "frais de vente": "selling fees",
    "tarifas de venta": "selling fees",
    "verkaufsgebühren": "selling fees",
    "commissioni di vendita": "selling fees",
    "verkoopkosten": "selling fees",
    "försäljningsavgifter": "selling fees",
    "opłaty za sprzedaż": "selling fees",
    "fba fees": "fba fees",
    "frais expédié par amazon": "fba fees",
    "frais pour le service expédié par amazon": "fba fees",
    "tarifas fba": "fba fees",
    "versand durch amazon gebühren": "fba fees",
    "gebühren zu versand durch amazon": "fba fees",
    "tariffe logistica di amazon": "fba fees",
    "costi del servizio logistica di amazon": "fba fees",
    "fba-kosten": "fba fees",
    "fba-avgifter": "fba fees",
    "opłaty za fba": "fba fees",
    "other transaction fees": "other transaction fees",
    "autres frais de transaction": "other transaction fees",
    "tarifas de otra transacción": "other transaction fees",
    "andere transaktionsgebühren": "other transaction fees",
    "altri costi relativi alle transazioni": "other transaction fees",
    "overige transactiekosten": "other transaction fees",
    "övriga transaktionsavgifter": "other transaction fees",
    "inne opłaty transakcyjne": "other transaction fees",
    "other": "other",
    "autre": "other",
    "autres": "other",
    "otro": "other",
    "andere": "other",
    "altro": "other",
    "overig": "other",
    "övrigt": "other",
    "inne": "other",
    "total": "total",
    "suma": "total",
    "summe": "total",
    "gesamt": "total",
    "totale": "total",
    "totaal": "total",
    "totalt": "total",
    "transaction status": "transaction status",
    "transaction release date": "transaction release date",
    "transaktionsstatus": "transaction status",
    "freigabedatum der transaktion": "transaction release date",
    "statut de la transaction": "transaction status",
    "date de sortie de la transaction": "transaction release date",
    "date de délivrance de la transaction": "transaction release date",
    "estado de la transacción": "transaction status",
    "fecha de liberación de la transacción": "transaction release date",
    "stato della transazione": "transaction status",
    "data di rilascio della transazione": "transaction release date",
}

LOW_SCHEMA = [
    "date/time", "settlement id", "type", "order id", "sku", "description", "quantity",
    "marketplace", "fulfillment", "order city", "order state", "order postal",
    "product sales", "shipping credits", "gift wrap credits", "promotional rebates",
    "sales tax collected", "marketplace facilitator tax", "selling fees", "fba fees",
    "other transaction fees", "other", "total",
]

HIGH_SCHEMA = [
    "date/time", "settlement id", "type", "order id", "sku", "description", "quantity",
    "marketplace", "fulfillment", "order city", "order state", "order postal",
    "tax collection model", "product sales", "product sales tax", "shipping credits",
    "shipping credits tax", "gift wrap credits", "gift wrap credits tax",
    "promotional rebates", "promotional rebates tax", "marketplace withheld tax",
    "selling fees", "fba fees", "other transaction fees", "other", "total",
]

HIGH_REG_SCHEMA = [
    "date/time", "settlement id", "type", "order id", "sku", "description", "quantity",
    "marketplace", "fulfillment", "order city", "order state", "order postal",
    "tax collection model", "product sales", "product sales tax", "shipping credits",
    "shipping credits tax", "gift wrap credits", "gift wrap credits tax",
    "regulatory fee", "tax on regulatory fee", "promotional rebates",
    "promotional rebates tax", "marketplace withheld tax", "selling fees", "fba fees",
    "other transaction fees", "other", "total",
]


def ascii_key(value):
    value = unicodedata.normalize("NFKD", value.lower())
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", value).strip()


def reference_key(country, raw_type, raw_description):
    return (
        clean_cell(country),
        ascii_key(clean_cell(raw_type)),
        ascii_key(clean_cell(raw_description)),
    )


def load_reference_rules():
    rules = {}
    if not REFERENCE_RULE_PATH.exists():
        return rules
    wb = load_workbook(REFERENCE_RULE_PATH, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = [clean_cell(str(v)) if v is not None else "" for v in next(rows)]
    index = {h: i for i, h in enumerate(headers)}
    required = {"国家", "type", "description", "记录类型"}
    if not required.issubset(index):
        wb.close()
        raise ValueError(f"Reference workbook missing columns: {required - set(index)}")
    for row in rows:
        country = clean_cell(row[index["国家"]] if row[index["国家"]] is not None else "")
        raw_type = clean_cell(row[index["type"]] if row[index["type"]] is not None else "")
        description = clean_cell(row[index["description"]] if row[index["description"]] is not None else "")
        record_type = clean_cell(row[index["记录类型"]] if row[index["记录类型"]] is not None else "")
        if country and raw_type and record_type:
            rules[reference_key(country, raw_type, description)] = record_type
    wb.close()
    return rules


REFERENCE_RULES = load_reference_rules()


def canonicalize_headers(headers):
    normalized = [norm_header(h) for h in headers]
    has_account = any(h == "account type" for h in normalized)
    has_status = any("status" in h or "statut" in h or "estado de la transacción" in h for h in normalized[-3:])
    has_regulatory = any("regulatory" in h or "reglamentaria" in h for h in normalized)
    has_tax_collection = any("tax collection model" in h or "perception" in h or "recaudación" in h for h in normalized)

    schema = None
    if len(headers) in (23, 25):
        schema = LOW_SCHEMA + (["transaction status", "transaction release date"] if len(headers) == 25 else [])
    elif has_account:
        schema = [
            "date/time", "settlement id", "type", "order id", "sku", "description", "quantity",
            "marketplace", "account type", "fulfillment", "order city", "order state", "order postal",
            "tax collection model", "product sales", "product sales tax", "shipping credits",
            "shipping credits tax", "gift wrap credits", "gift wrap credits tax",
        ]
        if has_regulatory:
            schema += ["regulatory fee", "tax on regulatory fee"]
        schema += [
            "promotional rebates", "promotional rebates tax", "marketplace withheld tax",
            "selling fees", "fba fees", "other transaction fees", "other", "total",
        ]
        if has_status or len(headers) == len(schema) + 2:
            schema += ["transaction status", "transaction release date"]
    elif has_tax_collection:
        schema = HIGH_REG_SCHEMA if has_regulatory else HIGH_SCHEMA
        if has_status or len(headers) == len(schema) + 2:
            schema += ["transaction status", "transaction release date"]

    result = []
    for idx, header in enumerate(headers):
        key = norm_header(header)
        canonical = CANONICAL_ALIASES.get(key)
        if not canonical and schema and idx < len(schema):
            canonical = schema[idx]
        result.append(canonical or clean_cell(header))
    return result


def parse_amount(value):
    raw = clean_cell(value)
    if raw == "":
        return ""
    if "," not in raw and all(ch in "0123456789.-" for ch in raw):
        return raw
    s = raw.replace("−", "-")
    negative = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^0-9,.\-]", "", s)
    if s in {"", "-", ".", ","}:
        return raw
    if negative:
        s = "-" + s.strip("()")
    if "," in s and "." in s:
        decimal_sep = "," if s.rfind(",") > s.rfind(".") else "."
        thousands_sep = "." if decimal_sep == "," else ","
        s = s.replace(thousands_sep, "")
        s = s.replace(decimal_sep, ".")
    elif "," in s:
        parts = s.split(",")
        if len(parts[-1]) in (1, 2):
            s = "".join(parts[:-1]) + "." + parts[-1]
        else:
            s = "".join(parts)
    elif "." in s:
        parts = s.split(".")
        if len(parts) > 2 or (len(parts[-1]) == 3 and len(parts[0]) > 0 and len(parts[0]) <= 3):
            s = "".join(parts)
    if not re.fullmatch(r"-?\d+(\.\d+)?", s):
        return raw
    if "." in s:
        left, right = s.split(".", 1)
        left = left.lstrip("0") or "0"
        right = right.rstrip("0")
        return f"{left}.{right}" if right else left
    return s.lstrip("0") or "0"


def parse_quantity(value):
    raw = clean_cell(value)
    if raw == "":
        return ""
    return parse_amount(raw)


def classify_record(country, raw_type, raw_description):
    reference_match = REFERENCE_RULES.get(reference_key(country, raw_type, raw_description))
    if reference_match:
        return reference_match

    t = ascii_key(clean_cell(raw_type))
    if any(x in t for x in ["order", "pedido", "commande", "bestellung", "ordine", "bestelling", "zamowienie", "bestallning"]):
        return "订单销售"
    if any(x in t for x in ["refund", "reembolso", "remboursement", "erstattung", "rimborso", "terugbetaling", "zwrot", "aterbetalning"]):
        return "退款"
    if any(x in t for x in ["debt", "deuda"]):
        return "信用卡扣款及其他债务回收"
    if any(x in t for x in ["transfer", "overboeking", "transferir", "uberweisung", "ubertrag", "trasferimento", "virement", "przelew", "overforing"]):
        return "资金转账"
    needs_description = any(x in t for x in [
        "service fee", "tarifa de prestacion", "frais de service", "servicegebuhr",
        "commissione per il servizio", "commissione di servizio", "servicekosten", "oplata", "serviceavgift",
        "fba inventory", "inventario", "stock", "lagerbestand", "lagergebuhr", "logistica di amazon", "magazyn", "lagring",
        "adjustment", "ajuste", "ajustement", "anpassung", "rettifica", "aanpassing",
        "korekta", "justering", "modifica", "gebuhrenerstattung",
    ])
    d = ascii_key(clean_cell(raw_description)) if needs_description else ""
    if needs_description and any(x in d for x in ["advertising", "publicidad", "publicite", "werbekosten", "pubblicita", "reklam"]):
        return "广告费"
    if needs_description and any(x in d for x in ["coupon", "cupon", "gutschein"]):
        return "优惠券费用"
    if needs_description and any(x in d for x in ["lightning", "deal", "oferta", "promocion", "promotion", "blitzangebot"]):
        return "促销活动费"
    if "fba inventory" in t or "inventario" in t or "stock" in t or "lagerbestand" in t or "lagergebuhr" in t or "logistica di amazon" in t or "magazyn" in t or "lagring" in t:
        if any(x in d for x in ["storage", "stockage", "almacenamiento", "lager", "stoccaggio", "magazyn", "lagring"]):
            return "FBA库存仓储费"
        if any(x in d for x in ["disposal", "removal", "elimin", "retrait", "entsorgung", "rimozione", "verwijder", "usun"]):
            return "FBA库存移除/弃置费"
        return "FBA库存费用"
    if "amazon fees" in t or "tarifas de amazon" in t or "frais amazon" in t or "amazon-gebuhren" in t or "amazon-kosten" in t or "gebuhren von amazon" in t or "gebuhrenerstattung" in t or "commissioni amazon" in t or "tariffe e commissioni amazon" in t or "oplaty amazon" in t:
        return "亚马逊平台费用"
    if any(x in t for x in ["service fee", "tarifa de prestacion", "frais de service", "servicegebuhr", "commissione per il servizio", "commissione di servizio", "servicekosten", "oplata za usluge", "serviceavgift"]):
        return "服务费"
    if any(x in t for x in ["adjustment", "ajuste", "ajustement", "anpassung", "rettifica", "aanpassing", "korekta", "justering", "modifica"]):
        if any(x in d for x in ["reimbursement", "remboursement", "reembolso", "erstattung", "rimborso", "vergoeding", "rekompensata"]):
            return "库存赔偿/调整"
        return "调整"
    if clean_cell(raw_type):
        return "其他"
    return "未识别"


AMOUNT_COLUMNS = {
    "product sales", "product sales tax", "shipping credits", "shipping credits tax",
    "gift wrap credits", "gift wrap credits tax", "regulatory fee", "tax on regulatory fee",
    "promotional rebates", "promotional rebates tax", "sales tax collected",
    "marketplace facilitator tax", "marketplace withheld tax", "selling fees", "fba fees",
    "other transaction fees", "other", "total",
}


def quarter_for(dt):
    return (dt.month - 1) // 3 + 1


def write_csv(path, headers, rows, chinese_row=True):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        if chinese_row:
            writer.writerow([CHINESE_HEADERS.get(h, h) for h in headers])
        for row in rows:
            writer.writerow([row.get(h, "") for h in headers])


def load_structured_files():
    files = sorted(
        p for p in ROOT.rglob("*.csv")
        if p.relative_to(ROOT).parts[0] not in SKIP_DIRS
    )
    loaded = []
    for path in files:
        print(f"loading {path.relative_to(ROOT)}", file=sys.stderr, flush=True)
        country = path.relative_to(ROOT).parts[0]
        raw_rows = rows_from_csv(path)
        header_idx = find_header(raw_rows)
        raw_headers = [clean_cell(c) for c in raw_rows[header_idx]]
        headers = canonicalize_headers(raw_headers)
        data_rows = [r for r in raw_rows[header_idx + 1:] if any(clean_cell(c) for c in r)]
        loaded.append({
            "country": country,
            "path": path,
            "header_row": header_idx + 1,
            "headers": headers,
            "data_rows": data_rows,
        })
    return loaded


def normalize_row(country, path, headers, row):
    record = {"country": country, "source file": str(path.relative_to(ROOT))}
    for idx, header in enumerate(headers):
        value = clean_cell(row[idx] if idx < len(row) else "")
        if header in AMOUNT_COLUMNS:
            value = parse_amount(value)
        elif header == "quantity":
            value = parse_quantity(value)
        record[header] = value

    dt = parse_date(record.get("date/time", ""))
    if dt:
        record["date"] = dt.date().isoformat()
        record["year"] = str(dt.year)
        record["month"] = str(dt.month)
    else:
        record["date"] = ""
        record["year"] = ""
        record["month"] = ""
    record["record type"] = classify_record(country, record.get("type", ""), record.get("description", ""))
    return record, dt


def main():
    if DETAIL_DIR.exists():
        shutil.rmtree(DETAIL_DIR)
    DETAIL_DIR.mkdir(parents=True, exist_ok=True)
    COUNTRY_DIR.mkdir(parents=True, exist_ok=True)
    if MERGED_QUARTERLY_DIR.exists():
        shutil.rmtree(MERGED_QUARTERLY_DIR)
    MERGED_QUARTERLY_DIR.mkdir(parents=True, exist_ok=True)

    loaded = load_structured_files()
    all_records = []
    country_records = defaultdict(list)
    country_schema_columns = defaultdict(set)
    detail_records = defaultdict(list)
    audit = {
        "source_files": [],
        "countries": {},
        "record_type_counts": {},
        "columns": {},
    }
    parse_failures = []

    for item in loaded:
        country = item["country"]
        path = item["path"]
        country_schema_columns[country].update(BASE_LEFT)
        country_schema_columns[country].update(item["headers"])
        print(f"processing {path.relative_to(ROOT)} rows={len(item['data_rows'])}", file=sys.stderr, flush=True)
        file_dates = []
        normalized_for_file = []
        for row in item["data_rows"]:
            record, dt = normalize_row(country, path, item["headers"], row)
            if dt is None:
                parse_failures.append(str(path.relative_to(ROOT)))
                continue
            file_dates.append(dt)
            normalized_for_file.append(record)
            key = (country, dt.year, quarter_for(dt))
            detail_records[key].append(record)
            country_records[country].append(record)
            all_records.append(record)

        audit["source_files"].append({
            "country": country,
            "file": str(path.relative_to(ROOT)),
            "header_row": item["header_row"],
            "source_rows": len(item["data_rows"]),
            "parsed_rows": len(normalized_for_file),
            "date_min": min(file_dates).date().isoformat() if file_dates else None,
            "date_max": max(file_dates).date().isoformat() if file_dates else None,
            "columns": item["headers"],
        })

    per_country_columns = {
        country: columns | {"country", "source file", "date", "year", "month", "record type"}
        for country, columns in country_schema_columns.items()
        if country in country_records
    }
    all_columns = set().union(*per_country_columns.values()) if per_country_columns else set()
    common_columns = set.intersection(*(cols for cols in per_country_columns.values() if cols)) if per_country_columns else set()
    fixed_left = [c for c in BASE_LEFT if c in all_columns and c != TOTAL_COL]
    ordered_common = [c for c in PREFERRED_ORDER if c in common_columns and c not in fixed_left and c != TOTAL_COL]
    ordered_diff = [c for c in PREFERRED_ORDER if c in all_columns and c not in common_columns and c not in fixed_left and c != TOTAL_COL]
    ordered_extra = sorted(c for c in all_columns if c not in fixed_left and c not in ordered_common and c not in ordered_diff and c != TOTAL_COL)
    final_columns = fixed_left + ordered_common + ordered_diff + ordered_extra
    if TOTAL_COL in all_columns:
        final_columns.append(TOTAL_COL)

    detail_headers = [c for c in final_columns if c not in {"country"}]
    for (country, year, quarter), records in sorted(detail_records.items()):
        path = DETAIL_DIR / f"{country}-{year}-Q{quarter}-交易明细.csv"
        write_csv(path, detail_headers, records, chinese_row=False)

    country_output_paths = {}
    country_columns = {}
    for country, records in sorted(country_records.items()):
        used = per_country_columns[country]
        country_cols = [c for c in final_columns if c in used or c in BASE_LEFT or c == TOTAL_COL]
        country_columns[country] = country_cols
        path = COUNTRY_DIR / f"{country}_清洗合并_2022-2026Q1.csv"
        write_csv(path, country_cols, records, chinese_row=True)
        country_output_paths[country] = str(path.relative_to(ROOT))

        dates = [datetime.fromisoformat(r["date"]) for r in records if r.get("date")]
        type_counts = Counter(r.get("record type", "") for r in records)
        audit["countries"][country] = {
            "rows": len(records),
            "date_min": min(dates).date().isoformat() if dates else None,
            "date_max": max(dates).date().isoformat() if dates else None,
            "expected_range": "2022-01-01 to 2026-03-31",
            "range_covers_expected": bool(dates and min(dates).date().isoformat() <= "2022-01-01" and max(dates).date().isoformat() >= "2026-03-31"),
            "record_type_counts": dict(type_counts),
            "output": country_output_paths[country],
        }

    total_path = OUTPUT_DIR / "亚马逊各国交易总表_2022-2026Q1.csv"
    write_csv(total_path, final_columns, all_records, chinese_row=True)

    merged_quarterly_files = []
    for country, records in sorted(country_records.items()):
        grouped = defaultdict(list)
        for record in records:
            if record.get("year") and record.get("month"):
                quarter = (int(record["month"]) - 1) // 3 + 1
                grouped[(int(record["year"]), quarter)].append(record)
        for (year, quarter), quarter_records in sorted(grouped.items()):
            path = MERGED_QUARTERLY_DIR / f"{country}-{year}-Q{quarter}-合并后备查.csv"
            write_csv(path, country_columns[country], quarter_records, chinese_row=True)
            merged_quarterly_files.append({
                "scope": "country",
                "country": country,
                "year": year,
                "quarter": quarter,
                "rows": len(quarter_records),
                "file": str(path.relative_to(ROOT)),
            })

    total_grouped = defaultdict(list)
    for record in all_records:
        if record.get("year") and record.get("month"):
            quarter = (int(record["month"]) - 1) // 3 + 1
            total_grouped[(int(record["year"]), quarter)].append(record)
    for (year, quarter), quarter_records in sorted(total_grouped.items()):
        path = MERGED_QUARTERLY_DIR / f"全部国家-{year}-Q{quarter}-合并后备查.csv"
        write_csv(path, final_columns, quarter_records, chinese_row=True)
        merged_quarterly_files.append({
            "scope": "all",
            "country": "全部国家",
            "year": year,
            "quarter": quarter,
            "rows": len(quarter_records),
            "file": str(path.relative_to(ROOT)),
        })

    audit["record_type_counts"] = dict(Counter(r.get("record type", "") for r in all_records))
    audit["columns"] = {
        "common_columns": ordered_common + ([TOTAL_COL] if TOTAL_COL in common_columns else []),
        "different_columns": ordered_diff + ordered_extra,
        "final_columns": final_columns,
    }
    audit["detail_backup_files"] = [
        {
            "country": country,
            "year": year,
            "quarter": quarter,
            "rows": len(records),
            "file": str((DETAIL_DIR / f"{country}-{year}-Q{quarter}-交易明细.csv").relative_to(ROOT)),
        }
        for (country, year, quarter), records in sorted(detail_records.items())
    ]
    audit["merged_quarterly_files"] = merged_quarterly_files
    audit["parse_failures_by_file"] = dict(Counter(parse_failures))
    audit["outputs"] = {
        "detail_backup_dir": str(DETAIL_DIR.relative_to(ROOT)),
        "country_merged_dir": str(COUNTRY_DIR.relative_to(ROOT)),
        "merged_quarterly_dir": str(MERGED_QUARTERLY_DIR.relative_to(ROOT)),
        "total": str(total_path.relative_to(ROOT)),
    }

    with (OUTPUT_DIR / "清洗审计报告.json").open("w", encoding="utf-8") as handle:
        json.dump(audit, handle, ensure_ascii=False, indent=2)

    print(json.dumps({
        "countries": audit["countries"],
        "total_rows": len(all_records),
        "total_file": str(total_path.relative_to(ROOT)),
        "detail_files": len(audit["detail_backup_files"]),
        "merged_quarterly_files": len(audit["merged_quarterly_files"]),
        "reference_rules": len(REFERENCE_RULES),
        "parse_failures_by_file": audit["parse_failures_by_file"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
