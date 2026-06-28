#!/usr/bin/env python3
import argparse
import csv
import json
import re
import shutil
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .process_amazon_reports import (
    AMOUNT_COLUMNS,
    BASE_LEFT,
    CANONICAL_ALIASES,
    CHINESE_HEADERS,
    PREFERRED_ORDER,
    TOTAL_COL,
    ascii_key,
    canonicalize_headers,
    classify_record,
    clean_cell,
    find_header,
    parse_amount,
    parse_date,
    parse_quantity,
    quarter_for,
)

PROJECT_ROOT = Path(__file__).resolve().parents[3]

COUNTRY_NAMES = [
    "美国", "加拿大", "英国", "德国", "法国", "意大利", "西班牙", "荷兰", "比利时",
    "爱尔兰", "波兰", "瑞典", "墨西哥", "沙特", "阿联酋", "日本", "澳洲", "澳大利亚",
]

JP_ALIASES = {
    "日付/時間": "date/time",
    "日付/時刻": "date/time",
    "決済番号": "settlement id",
    "トランザクションの種類": "type",
    "注文番号": "order id",
    "sku": "sku",
    "説明": "description",
    "数量": "quantity",
    "amazon 出品サービス": "marketplace",
    "フルフィルメント": "fulfillment",
    "市町村": "order city",
    "都道府県": "order state",
    "郵便番号": "order postal",
    "税金徴収型": "tax collection model",
    "商品売上": "product sales",
    "商品の売上税": "product sales tax",
    "配送料": "shipping credits",
    "配送料の税金": "shipping credits tax",
    "ギフト包装手数料": "gift wrap credits",
    "ギフト包装クレジットの税金": "gift wrap credits tax",
    "プロモーション割引額": "promotional rebates",
    "プロモーション割引の税金": "promotional rebates tax",
    "Amazonポイントの費用": "cost of points granted",
    "源泉徴収税を伴うマーケットプレイス": "marketplace withheld tax",
    "手数料": "selling fees",
    "FBA 手数料": "fba fees",
    "トランザクションに関するその他の手数料": "other transaction fees",
    "その他": "other",
    "合計": "total",
    "トランザクションのステータス": "transaction status",
    "トランザクション開始日": "transaction release date",
}

EXTRA_ALIASES = {
    "settlement id": "settlement id",
    "settlement id": "settlement id",
    "id transazione": "settlement id",
    **{k.lower(): v for k, v in JP_ALIASES.items()},
}

EXTRA_CHINESE_HEADERS = {
    "brand": "品牌",
    "source format": "源文件格式",
    "sheet name": "工作表",
    "header row": "表头行",
    "quarter": "季度",
    "currency": "货币",
    "country inference": "国家识别方式",
    "cost of points granted": "积分授予成本",
}

for key, value in EXTRA_CHINESE_HEADERS.items():
    CHINESE_HEADERS.setdefault(key, value)

AMOUNT_COLUMNS = set(AMOUNT_COLUMNS) | {"cost of points granted"}


def norm(value):
    value = clean_cell(value).lower()
    value = re.sub(r"\s+", " ", value)
    return value.strip(" :：")


def detect_country_and_brand(path: Path, input_dir: Path):
    text = str(path.relative_to(input_dir))
    country = next((c for c in COUNTRY_NAMES if c in text), "")
    parts = re.split(r"[-_]+", path.stem)
    parts = [p for p in parts if p and not re.fullmatch(r"\d{6}|\d{4}", p)]
    stop = {"payment", "Payment", "报告", "交易报告", "按季度统计"}
    brand = next((p for p in parts if p not in stop and p not in COUNTRY_NAMES), "")
    method = "filename" if country else "unresolved"
    return brand, country, method


def read_csv_rows(path: Path):
    encodings = ["utf-8-sig", "utf-16", "gb18030"]
    last_error = None
    for enc in encodings:
        try:
            text = path.read_text(encoding=enc)
            sample = text[:4096]
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            return list(csv.reader(text.splitlines(), dialect))
        except Exception as exc:
            last_error = exc
    raise ValueError(f"无法读取 CSV: {path}: {last_error}")


def read_xlsx_sheets(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([("" if v is None else str(v)) for v in row])
            yield ws.title, rows
    finally:
        wb.close()


def find_header_extended(rows):
    try:
        return find_header(rows)
    except ValueError:
        for i, row in enumerate(rows[:30]):
            normalized = [norm(c) for c in row]
            if any(h in normalized for h in ("日付/時間", "日付/時刻")) and "トランザクションの種類" in normalized:
                return i
        raise


def canonicalize_headers_extended(headers):
    first = canonicalize_headers(headers)
    result = []
    for raw, canonical in zip(headers, first):
        key = norm(raw)
        result.append(EXTRA_ALIASES.get(key, canonical))
    return result


def repair_row_width(headers, row):
    if len(row) <= len(headers):
        return row
    try:
        desc_idx = headers.index("description")
    except ValueError:
        return row
    extra = len(row) - len(headers)
    repaired = (
        list(row[:desc_idx])
        + [",".join(clean_cell(c) for c in row[desc_idx:desc_idx + extra + 1])]
        + list(row[desc_idx + extra + 1:])
    )
    return repaired


def parse_date_extended(value):
    dt = parse_date(value)
    if dt:
        return dt
    raw = clean_cell(value)
    if not raw:
        return None
    raw_no_tz = re.sub(r"\s+(JST|UTC|GMT[+-]?\d*|PDT|PST|BST)$", "", raw, flags=re.IGNORECASE)
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw_no_tz[:19], fmt)
        except ValueError:
            pass
    return None


def classify_record_extended(country, raw_type, raw_description):
    raw = clean_cell(raw_type)
    raw_lower = raw.lower()
    desc_raw = clean_cell(raw_description).lower()
    folded = ascii_key(raw)
    if any(x in folded for x in ["skuld", "schuld", "debt", "deuda", "dette", "debito"]):
        return "信用卡扣款及其他债务回收"
    desc = ascii_key(clean_cell(raw_description))
    if any(x in folded for x in ["liquidation", "liquidationen"]):
        return "FBA清仓/清算"
    if any(x in folded for x in [
        "fba transaction fees",
        "transaktionsgebuhren fur versand durch amazon",
        "tarifas de transaccion de logistica de amazon",
        "frais de transaction expedie par amazon",
        "fulfilment by amazon (fba) transaction fees",
    ]):
        if any(x in desc for x in ["storage", "lager", "almacenamiento", "stockage", "long-term"]):
            return "FBA库存仓储费"
        if "customer returns fee" in desc:
            return "FBA退货处理费"
        return "FBA交易费"
    if "fulfilment by amazon inventory fee" in folded:
        if any(x in desc for x in ["disposal", "removal"]):
            return "FBA库存移除/弃置费"
        if any(x in desc for x in ["storage", "long-term"]):
            return "FBA库存仓储费"
        return "FBA库存费用"
    if any(x in folded for x in ["trasferir", "transferir"]):
        return "资金转账"
    if any(x in folded for x in ["tarifa de servicio", "opata za usuge", "oplata za usluge"]) or "opłata za usługę" in raw_lower:
        if any(x in desc for x in ["advertising", "publicidad", "publicite", "werbung", "reklam", "koszt reklamy"]) or "koszt reklamy" in desc_raw:
            return "广告费"
        return "服务费"
    if "注文" in raw:
        return "订单销售"
    if "返金" in raw or "返金" in clean_cell(raw_description):
        return "退款"
    if any(x in raw for x in ["振込", "振替", "送金"]):
        return "资金转账"
    if any(x in raw for x in ["サービス料", "手数料"]):
        desc = clean_cell(raw_description)
        if "広告" in desc:
            return "广告费"
        return "服务费"
    return classify_record(country, raw_type, raw_description)


def is_note_row(row, header_count):
    cells = [clean_cell(c) for c in row[:header_count]]
    nonempty = [c for c in cells if c]
    if not nonempty:
        return True
    normalized = [norm(c) for c in cells]
    if (
        any(h in normalized for h in ("date/time", "date/heure", "fecha/hora", "fecha y hora", "datum/uhrzeit", "data/ora", "datum/tijd", "日付/時間", "日付/時刻"))
        and any(h in normalized for h in ("type", "tipo", "typ", "トランザクションの種類"))
    ):
        return True
    if len(nonempty) <= 2 and parse_date_extended(cells[0]) is None:
        return True
    return False


def infer_currency(rows):
    for row in rows[:10]:
        text = " ".join(clean_cell(c) for c in row if c)
        m = re.search(r"\b(USD|CAD|GBP|EUR|PLN|SEK|MXN|AUD|JPY|SAR|AED)\b", text)
        if m:
            return m.group(1)
        if "Euro" in text:
            return "EUR"
        if "単位は円" in text:
            return "JPY"
        if "local currency" in text:
            return "LOCAL"
    return ""


def file_entries(input_dir: Path):
    for path in sorted(input_dir.rglob("*")):
        if path.suffix.lower() == ".csv":
            yield path, "csv"
        elif path.suffix.lower() == ".xlsx":
            yield path, "xlsx"
        elif path.suffix.lower() == ".xls":
            yield path, "xls"


def display_path(path: Path, base: Path | None = None):
    base = base or PROJECT_ROOT
    try:
        return str(path.relative_to(base))
    except ValueError:
        return str(path)


def load_sources(input_dir: Path):
    loaded = []
    unresolved = []
    unsupported = []
    for path, fmt in file_entries(input_dir):
        if fmt == "xls":
            unsupported.append(str(path.relative_to(input_dir)))
            continue
        brand, country, country_method = detect_country_and_brand(path, input_dir)
        if not country:
            unresolved.append(str(path.relative_to(input_dir)))
        sheets = [(path.stem, read_csv_rows(path))] if fmt == "csv" else list(read_xlsx_sheets(path))
        for sheet_name, rows in sheets:
            try:
                header_idx = find_header_extended(rows)
            except Exception as exc:
                raise ValueError(f"表头识别失败: {path.relative_to(input_dir)} / {sheet_name}: {exc}") from exc
            raw_headers = [clean_cell(c) for c in rows[header_idx]]
            headers = canonicalize_headers_extended(raw_headers)
            data_rows = [
                repair_row_width(headers, r)
                for r in rows[header_idx + 1:]
                if not is_note_row(r, len(headers))
            ]
            loaded.append({
                "path": path,
                "format": fmt,
                "sheet_name": sheet_name,
                "brand": brand,
                "country": country,
                "country_method": country_method,
                "header_row": header_idx + 1,
                "raw_headers": raw_headers,
                "headers": headers,
                "rows": rows,
                "data_rows": data_rows,
                "currency": infer_currency(rows),
            })
    return loaded, unresolved, unsupported


def normalize_row(item, row):
    country = item["country"]
    record = {
        "country": country,
        "brand": item["brand"],
        "source file": display_path(item["path"]),
        "source format": item["format"],
        "sheet name": item["sheet_name"],
        "header row": item["header_row"],
        "currency": item["currency"],
        "country inference": item["country_method"],
    }
    amount_failures = []
    for idx, header in enumerate(item["headers"]):
        raw_value = clean_cell(row[idx] if idx < len(row) else "")
        if header in AMOUNT_COLUMNS:
            value = parse_amount(raw_value)
            if raw_value and value == raw_value and not re.fullmatch(r"-?\d+(\.\d+)?", value):
                amount_failures.append((header, raw_value))
        elif header == "quantity":
            value = parse_quantity(raw_value)
        else:
            value = raw_value
        record[header] = value

    dt = parse_date_extended(record.get("date/time", ""))
    if dt:
        record["date"] = dt.date().isoformat()
        record["year"] = str(dt.year)
        record["month"] = str(dt.month)
        record["quarter"] = f"Q{quarter_for(dt)}"
    else:
        record["date"] = ""
        record["year"] = ""
        record["month"] = ""
        record["quarter"] = ""
    record["record type"] = classify_record_extended(country, record.get("type", ""), record.get("description", ""))
    return record, dt, amount_failures


def ordered_columns(per_country_columns):
    all_columns = set().union(*per_country_columns.values()) if per_country_columns else set()
    common_columns = set.intersection(*(cols for cols in per_country_columns.values() if cols)) if per_country_columns else set()
    left = ["country", "brand", "source file", "source format", "sheet name", "header row", "currency", "country inference"]
    left += [c for c in BASE_LEFT if c not in {"country", "source file"}]
    left += ["quarter"]
    fixed_left = [c for c in left if c in all_columns and c != TOTAL_COL]
    common = [c for c in PREFERRED_ORDER if c in common_columns and c not in fixed_left and c != TOTAL_COL]
    diff = [c for c in PREFERRED_ORDER if c in all_columns and c not in common_columns and c not in fixed_left and c != TOTAL_COL]
    extra = sorted(c for c in all_columns if c not in fixed_left and c not in common and c not in diff and c != TOTAL_COL)
    result = fixed_left + common + diff + extra
    if TOTAL_COL in all_columns:
        result.append(TOTAL_COL)
    return result, common, diff + extra


def typed_value(header, value):
    if value == "":
        return ""
    if header in {"year", "month", "header row"}:
        try:
            return int(value)
        except Exception:
            return value
    if header == "date":
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except Exception:
            return value
    if header in AMOUNT_COLUMNS or header == "quantity":
        try:
            return float(value) if "." in str(value) else int(value)
        except Exception:
            return value
    return value


def setup_sheet(ws, freeze="A3"):
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 2)):
        for cell in row:
            cell.font = Font(bold=True, color="FFFFFF" if cell.row == 1 else "000000")
            cell.fill = PatternFill("solid", fgColor="1F4E78" if cell.row == 1 else "D9EAF7")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 8), len(value) + 2), 42)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_table_xlsx(path: Path, sheet_name: str, columns, rows, include_chinese=True):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(columns)
    if include_chinese:
        ws.append([CHINESE_HEADERS.get(c, c) for c in columns])
    for row in rows:
        ws.append([typed_value(c, row.get(c, "")) for c in columns])
    setup_sheet(ws)
    wb.save(path)


def write_audit_xlsx(path: Path, audit):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ["指标", "值"],
        ["源文件总数", audit["source_file_count"]],
        ["CSV文件数", audit["source_file_formats"].get("csv", 0)],
        ["XLSX文件数", audit["source_file_formats"].get("xlsx", 0)],
        ["XLS文件数", audit["source_file_formats"].get("xls", 0)],
        ["有效源明细行数", audit["source_valid_rows"]],
        ["成功解析行数", audit["parsed_rows"]],
        ["总表行数", audit["total_rows"]],
        ["日期解析失败数", audit["date_parse_failure_count"]],
        ["金额清洗异常数", audit["amount_failure_count"]],
        ["无法识别国家文件数", len(audit["unresolved_country_files"])],
    ]
    for row in rows:
        ws.append(row)
    setup_sheet(ws, "A2")

    ws_files = wb.create_sheet("Files")
    file_cols = ["file", "format", "sheet", "brand", "country", "country_method", "header_row", "source_rows", "parsed_rows", "date_min", "date_max", "field_count", "currency"]
    ws_files.append(file_cols)
    for item in audit["files"]:
        ws_files.append([item.get(c, "") for c in file_cols])
    setup_sheet(ws_files, "A2")

    ws_country = wb.create_sheet("Country_Rows")
    ws_country.append(["国家", "行数", "日期最小值", "日期最大值"])
    for country, item in sorted(audit["countries"].items()):
        ws_country.append([country, item["rows"], item["date_min"], item["date_max"]])
    setup_sheet(ws_country, "A2")

    ws_types = wb.create_sheet("Record_Types")
    ws_types.append(["国家", "记录类型", "行数"])
    for country, counts in sorted(audit["record_type_by_country"].items()):
        for record_type, count in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
            ws_types.append([country, record_type, count])
    setup_sheet(ws_types, "A2")

    ws_fail = wb.create_sheet("Failures")
    ws_fail.append(["类型", "文件", "行号/字段", "值"])
    for item in audit["date_parse_failures"]:
        ws_fail.append(["日期解析失败", item["file"], item["row_number"], item["date_time"]])
    for item in audit["amount_failures"]:
        ws_fail.append(["金额清洗异常", item["file"], item["field"], item["value"]])
    for item in audit["unresolved_country_files"]:
        ws_fail.append(["国家未识别", item, "", ""])
    for item in audit["unsupported_files"]:
        ws_fail.append(["暂不支持格式", item, "", ""])
    setup_sheet(ws_fail, "A2")

    ws_cols = wb.create_sheet("Fields")
    ws_cols.append(["字段", "中文翻译", "出现国家数", "出现国家"])
    for field, countries in sorted(audit["field_coverage"].items()):
        ws_cols.append([field, CHINESE_HEADERS.get(field, field), len(countries), "、".join(sorted(countries))])
    setup_sheet(ws_cols, "A2")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main(argv=None):
    parser = argparse.ArgumentParser(description="Clean and merge Amazon payment transaction detail files.")
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--output-dir", default="outputs/202605payment_cleaned")
    parser.add_argument("--label", default=None)
    parser.add_argument(
        "--include-country-files",
        action="store_true",
        help="额外生成各国家清洗合并表。默认只生成总表和审计报告。",
    )
    parser.add_argument(
        "--include-quarter-backup",
        action="store_true",
        help="额外生成清洗后季度备查表。默认只生成总表和审计报告。",
    )
    args = parser.parse_args(argv)

    input_dir = (PROJECT_ROOT / args.input_dir).resolve() if not Path(args.input_dir).is_absolute() else Path(args.input_dir)
    output_dir = (PROJECT_ROOT / args.output_dir).resolve() if not Path(args.output_dir).is_absolute() else Path(args.output_dir)
    label = args.label or input_dir.name

    if output_dir.exists():
        shutil.rmtree(output_dir)
    country_dir = output_dir / "country_merged"
    quarter_dir = output_dir / "merged_quarterly_backup"

    loaded, unresolved, unsupported = load_sources(input_dir)
    all_records = []
    country_records = defaultdict(list)
    per_country_columns = defaultdict(set)
    field_coverage = defaultdict(set)
    date_parse_failures = []
    amount_failures = []
    files_audit = []

    for item in loaded:
        parsed_rows = []
        dates = []
        for idx, row in enumerate(item["data_rows"], start=item["header_row"] + 1):
            record, dt, row_amount_failures = normalize_row(item, row)
            if dt is None:
                date_parse_failures.append({
                    "file": display_path(item["path"]),
                    "row_number": idx,
                    "date_time": record.get("date/time", ""),
                })
                continue
            if row_amount_failures:
                for field, value in row_amount_failures:
                    amount_failures.append({
                        "file": display_path(item["path"]),
                        "field": field,
                        "value": value,
                    })
            parsed_rows.append(record)
            dates.append(dt)
            all_records.append(record)
            country_records[item["country"]].append(record)
            per_country_columns[item["country"]].update(record.keys())
            for key in record:
                field_coverage[key].add(item["country"])
        files_audit.append({
            "file": display_path(item["path"]),
            "format": item["format"],
            "sheet": item["sheet_name"],
            "brand": item["brand"],
            "country": item["country"],
            "country_method": item["country_method"],
            "header_row": item["header_row"],
            "source_rows": len(item["data_rows"]),
            "parsed_rows": len(parsed_rows),
            "date_min": min(dates).date().isoformat() if dates else "",
            "date_max": max(dates).date().isoformat() if dates else "",
            "field_count": len(item["headers"]),
            "currency": item["currency"],
        })

    final_columns, common_columns, diff_columns = ordered_columns(per_country_columns)
    country_files = []
    if args.include_country_files:
        for country, records in sorted(country_records.items()):
            cols = [c for c in final_columns if c in per_country_columns[country] or c in {"country", "brand", "source file", "date", "year", "month", "quarter", "record type"}]
            path = country_dir / f"{country}_清洗合并_{label}.xlsx"
            write_table_xlsx(path, "明细", cols, records)
            country_files.append({"country": country, "rows": len(records), "file": display_path(path)})

    total_path = output_dir / f"亚马逊各国交易总表_{label}.xlsx"
    write_table_xlsx(total_path, "All", final_columns, all_records)

    quarter_files = []
    if args.include_quarter_backup:
        by_country_quarter = defaultdict(list)
        all_by_quarter = defaultdict(list)
        for record in all_records:
            key = (record["country"], int(record["year"]), record["quarter"])
            by_country_quarter[key].append(record)
            all_by_quarter[(int(record["year"]), record["quarter"])].append(record)
        for (country, year, quarter), rows in sorted(by_country_quarter.items()):
            cols = [c for c in final_columns if c in per_country_columns[country] or c in {"country", "brand", "source file", "date", "year", "month", "quarter", "record type"}]
            path = quarter_dir / f"{country}-{year}-{quarter}.xlsx"
            write_table_xlsx(path, "明细", cols, rows)
            quarter_files.append({"scope": "country", "country": country, "year": year, "quarter": quarter, "rows": len(rows), "file": display_path(path)})
        for (year, quarter), rows in sorted(all_by_quarter.items()):
            path = quarter_dir / f"全部国家-{year}-{quarter}.xlsx"
            write_table_xlsx(path, "明细", final_columns, rows)
            quarter_files.append({"scope": "all", "country": "全部国家", "year": year, "quarter": quarter, "rows": len(rows), "file": display_path(path)})

    country_summary = {}
    record_type_by_country = {}
    for country, records in sorted(country_records.items()):
        dates = [r["date"] for r in records if r.get("date")]
        country_summary[country] = {
            "rows": len(records),
            "date_min": min(dates) if dates else "",
            "date_max": max(dates) if dates else "",
        }
        record_type_by_country[country] = dict(Counter(r.get("record type", "") for r in records))

    audit = {
        "input_dir": display_path(input_dir),
        "output_dir": display_path(output_dir),
        "source_file_count": len(list(file_entries(input_dir))),
        "source_file_formats": dict(Counter(fmt for _path, fmt in file_entries(input_dir))),
        "source_valid_rows": sum(item["source_rows"] for item in files_audit),
        "parsed_rows": len(all_records),
        "total_rows": len(all_records),
        "date_parse_failure_count": len(date_parse_failures),
        "amount_failure_count": len(amount_failures),
        "date_parse_failures": date_parse_failures,
        "amount_failures": amount_failures,
        "unresolved_country_files": unresolved,
        "unsupported_files": unsupported,
        "files": files_audit,
        "countries": country_summary,
        "record_type_by_country": record_type_by_country,
        "record_type_counts": dict(Counter(r.get("record type", "") for r in all_records)),
        "field_coverage": {k: sorted(v) for k, v in field_coverage.items()},
        "common_columns": common_columns,
        "different_columns": diff_columns,
        "final_columns": final_columns,
        "country_files": country_files,
        "quarter_files": quarter_files,
        "outputs": {
            "total": display_path(total_path),
            "country_dir": display_path(country_dir) if args.include_country_files else "",
            "quarter_dir": display_path(quarter_dir) if args.include_quarter_backup else "",
            "audit_json": display_path(output_dir / "清洗审计报告.json"),
            "audit_xlsx": display_path(output_dir / "清洗审计报告.xlsx"),
        },
    }
    (output_dir / "清洗审计报告.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    write_audit_xlsx(output_dir / "清洗审计报告.xlsx", audit)
    print(json.dumps({
        "source_files": audit["source_file_count"],
        "formats": audit["source_file_formats"],
        "total_rows": audit["total_rows"],
        "countries": {k: v["rows"] for k, v in country_summary.items()},
        "date_parse_failures": len(date_parse_failures),
        "amount_failures": len(amount_failures),
        "country_files": len(country_files),
        "quarter_files": len(quarter_files),
        "total": audit["outputs"]["total"],
        "audit": audit["outputs"]["audit_xlsx"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
