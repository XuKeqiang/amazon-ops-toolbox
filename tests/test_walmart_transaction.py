import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.ops_toolbox.walmart_transaction.batch import process_walmart_transaction_folder


class WalmartTransactionProcessingTest(unittest.TestCase):
    def test_process_folder_maps_by_header_and_writes_outputs(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "沃尔玛账单明细新版（20260502-20260516）.xlsx"
            _write_source_workbook(source)
            (root / ".DS_Store").write_text("skip", encoding="utf-8")

            job = process_walmart_transaction_folder(source.parent, root / "outputs", "job001")

            self.assertEqual(job.summary["source_files"], 1)
            self.assertEqual(job.summary["skipped_files"], 1)
            self.assertEqual(job.summary["output_rows"], 1)
            self.assertTrue(job.total_path.exists())
            self.assertTrue(job.audit_path.exists())

            ws = load_workbook(job.total_path, read_only=True, data_only=True)["账单明细"]
            headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            row = [cell for cell in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))]
            values = dict(zip(headers, row))

            self.assertEqual(values["Source Period"], "20260502-20260516")
            self.assertEqual(values["Year"], 2026)
            self.assertEqual(values["Month"], 5)
            self.assertEqual(values["Quarter"], "Q2")
            self.assertEqual(values["Transaction Type(中文)"], "销售")
            self.assertEqual(values["Amount Type(中文)"], "商品售价")
            self.assertEqual(values["Data Type Description"], "商品售价")
            self.assertEqual(values["Partner Item Id"], "PID-001")
            self.assertEqual(values["Partner GTIN"], "GTIN-001")
            self.assertEqual(values["Partner Item Name"], "Test Product")
            self.assertEqual(values["Fulfillment Type"], "Walmart-fulfilled(WFS)")

    def test_process_folder_infers_period_from_content_when_filename_lacks_it(self):
        # 文件名不含 20YYYYMMDD-20YYYYMMDD 账期、但内容有交易日期时，
        # 应回退由内容推断账期，且不误报「文件名未识别来源账期」。
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "沃尔玛账单明细-Bikoney.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "report"
            ws.append([
                "Period Start Date", "Period End Date", "Total Payable", "Currency",
                "Transaction Key", "Transaction Posted Timestamp", "Transaction Type",
                "Transaction Description", "Customer Order #", "Customer Order line #",
                "Purchase Order #", "Purchase Order line #", "Amount", "Amount Type",
                "Ship Qty", "Commission Rate", "Base Commission Rate",
                "Transaction Reason Description", "Partner Item Id", "Partner GTIN",
                "Partner Item Name", "Product Tax Code", "Ship to State", "Ship to City",
                "Ship to Zipcode", "Contract Category", "Product Type", "Commission Rule",
                "Shipping Method", "Fulfillment Type", "Fulfillment Details",
                "Original Commission", "Commission Incentive Program", "Commission Saving",
                "Customer Promo Type", "Total Walmart Funded Savings Program", "Campaign Id",
                "Item Condition",
            ])
            ws.append(["Number of Lines in file 1"])
            ws.append([None, None, None, None, "K1", "05/10/2026", "Sale", "Purchase",
                       "ORDER-001", 1, "PO-001", 1, 99.99, "Product Price", 1, None, None,
                       None, "PID-001", "GTIN-001", "Test Product", "TAX", "CA",
                       "Los Angeles", "90001", "Furniture & Decor", "End Tables", "Rule",
                       "Marketplace standard", "Walmart-fulfilled(WFS)", "Delivery", None,
                       None, None, None, None, None, None, None])
            ws.append([None, None, None, None, "K2", "05/15/2026", "Sale", "Purchase",
                       "ORDER-002", 1, "PO-002", 1, 10.00, "Product Price", 1, None, None,
                       None, "PID-002", "GTIN-002", "Test Product 2", "TAX", "CA",
                       "Los Angeles", "90001", "Furniture & Decor", "End Tables", "Rule",
                       "Marketplace standard", "Walmart-fulfilled(WFS)", "Delivery", None,
                       None, None, None, None, None, None, None])
            wb.save(source)

            job = process_walmart_transaction_folder(source.parent, root / "outputs", "job002")

            self.assertEqual(job.summary["warnings"], 0)
            file_row = job.rows[0]
            self.assertEqual(file_row["status"], "通过")
            self.assertEqual(file_row["source_period"], "20260510-20260515")
            ws = load_workbook(job.total_path, read_only=True, data_only=True)["账单明细"]
            first_row = [cell for cell in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))]
            headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            self.assertEqual(dict(zip(headers, first_row))["Source Period"], "20260510-20260515")


def _write_source_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "report"
    ws.append([
        "Period Start Date",
        "Period End Date",
        "Total Payable",
        "Currency",
        "Transaction Key",
        "Transaction Posted Timestamp",
        "Transaction Type",
        "Transaction Description",
        "Customer Order #",
        "Customer Order line #",
        "Purchase Order #",
        "Purchase Order line #",
        "Amount",
        "Amount Type",
        "Ship Qty",
        "Commission Rate",
        "Base Commission Rate",
        "Transaction Reason Description",
        "Partner Item Id",
        "Partner GTIN",
        "Partner Item Name",
        "Product Tax Code",
        "Ship to State",
        "Ship to City",
        "Ship to Zipcode",
        "Contract Category",
        "Product Type",
        "Commission Rule",
        "Shipping Method",
        "Fulfillment Type",
        "Fulfillment Details",
        "Original Commission",
        "Commission Incentive Program",
        "Commission Saving",
        "Customer Promo Type",
        "Total Walmart Funded Savings Program",
        "Campaign Id",
        "Item Condition",
    ])
    ws.append(["Number of Lines in file 1"])
    ws.append([
        None,
        None,
        None,
        None,
        "2026_05_10_1",
        "05/10/2026",
        "Sale",
        "Purchase",
        "ORDER-001",
        1,
        "PO-001",
        1,
        99.99,
        "Product Price",
        1,
        None,
        None,
        None,
        "PID-001",
        "GTIN-001",
        "Test Product",
        "TAX",
        "CA",
        "Los Angeles",
        "90001",
        "Furniture & Decor",
        "End Tables",
        "Rule",
        "Marketplace standard",
        "Walmart-fulfilled(WFS)",
        "Delivery",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ])
    wb.save(path)


if __name__ == "__main__":
    unittest.main()
